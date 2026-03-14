"""
Telegram TeraBox Link Extractor — connects to Telegram, scans a specific
channel/group for TeraBox links, saves them to Excel, then auto-triggers
the batch downloader.

First run will prompt for:
  - api_id and api_hash (from https://my.telegram.org → "API development tools")
  - Phone number + OTP for Telegram login
  - Session is saved to telegram.session for future runs

Usage:
  python telegram_extractor.py --channel @mychannel               # scan last 100 messages
  python telegram_extractor.py --channel @mychannel --limit 500   # scan last 500 messages
  python telegram_extractor.py --channel https://t.me/mychannel   # t.me link also works
  python telegram_extractor.py --channel @mychannel --no-download # extract only, skip download

Credentials are cached in config.json (already in .gitignore).
"""

import argparse
import asyncio
import os
import re
import sys
import logging
from datetime import datetime
from urllib.parse import urlparse

from telethon import TelegramClient
from telethon.tl.types import MessageEntityUrl, MessageEntityTextUrl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Import project config ───────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import SUPPORTED_DOMAINS, load_config, save_config

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
SESSION_FILE = os.path.join(PROJECT_DIR, "telegram")  # creates telegram.session
HEADERS = ["Link", "Status", "Filename", "Size (MB)", "Date"]
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Regex to find TeraBox URLs in message text
TERABOX_URL_PATTERN = re.compile(
    r'https?://(?:www\.)?(?:' +
    '|'.join(re.escape(d.replace("www.", "")) for d in SUPPORTED_DOMAINS if not d.startswith("www.")) +
    r')[^\s<>"\')\]]*',
    re.IGNORECASE,
)


def get_telegram_credentials() -> tuple[int, str]:
    """
    Get Telegram API credentials from config.json.
    Prompts user interactively if not configured.
    """
    cfg = load_config()
    api_id = cfg.get("telegram_api_id")
    api_hash = cfg.get("telegram_api_hash")

    if api_id and api_hash:
        return int(api_id), str(api_hash)

    print("\n" + "=" * 60)
    print("  TELEGRAM API SETUP")
    print("=" * 60)
    print()
    print("  You need API credentials from Telegram:")
    print("  1. Go to: https://my.telegram.org")
    print("  2. Log in with your phone number")
    print("  3. Click 'API development tools'")
    print("  4. Create an app (any name/short name)")
    print("  5. Copy the api_id and api_hash")
    print()

    while True:
        try:
            api_id_input = input("  Enter api_id (number): ").strip()
            api_id = int(api_id_input)
            break
        except ValueError:
            print("  ⚠ api_id must be a number. Try again.")

    api_hash = input("  Enter api_hash (string): ").strip()
    if not api_hash:
        print("  ⚠ api_hash cannot be empty.")
        sys.exit(1)

    # Save to config.json
    cfg["telegram_api_id"] = api_id
    cfg["telegram_api_hash"] = api_hash
    save_config(cfg)
    print(f"\n  Credentials saved to config.json")
    print()

    return api_id, api_hash


def extract_urls_from_message(message) -> list[str]:
    """Extract TeraBox URLs from a Telegram message (text + entities)."""
    urls = set()

    # 1. Extract from message text using regex
    if message.text:
        found = TERABOX_URL_PATTERN.findall(message.text)
        urls.update(found)

    # 2. Extract from message entities (hyperlinks)
    if message.entities:
        for entity in message.entities:
            if isinstance(entity, MessageEntityTextUrl):
                url = entity.url
                if is_terabox_domain(url):
                    urls.add(url)
            elif isinstance(entity, MessageEntityUrl) and message.text:
                url = message.text[entity.offset: entity.offset + entity.length]
                if is_terabox_domain(url):
                    if not url.startswith("http"):
                        url = "https://" + url
                    urls.add(url)

    # 3. Check for web page preview
    if message.web_preview and hasattr(message.web_preview, 'url'):
        url = message.web_preview.url
        if is_terabox_domain(url):
            urls.add(url)

    return list(urls)


def is_terabox_domain(url: str) -> bool:
    """Check if a URL belongs to a known TeraBox domain."""
    try:
        host = urlparse(url).hostname
        return host in SUPPORTED_DOMAINS if host else False
    except Exception:
        return False


def normalize_channel(channel: str) -> str:
    """Normalize channel input: handle @username, t.me links, invite links."""
    channel = channel.strip()

    # Handle t.me links: https://t.me/channelname
    match = re.match(r'https?://t\.me/([^/?]+)', channel)
    if match:
        return match.group(1)

    # Handle t.me/+invite links (private channels)
    if 't.me/+' in channel or 't.me/joinchat/' in channel:
        return channel  # telethon handles invite links directly

    # Remove @ prefix if present (telethon accepts both)
    return channel


def load_existing_links(filepath: str) -> set[str]:
    """Load existing links from Excel to deduplicate."""
    existing = set()
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    existing.add(str(row[0]).strip())
            wb.close()
        except Exception as e:
            logger.warning(f"Could not read existing Excel: {e}")
    return existing


def save_to_excel(links: list[dict], filepath: str, append: bool = True):
    """
    Save extracted links to Excel file.
    Each link dict: {url: str, date: datetime, message_id: int}
    """
    if append and os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "TeraBox Links"

        # Write headers
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 65
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 22
        start_row = 2

    # Write links
    for i, link_info in enumerate(links):
        row = start_row + i
        ws.cell(row=row, column=1, value=link_info["url"])
        ws.cell(row=row, column=2, value="")  # Status — empty, to be filled by batch_downloader
        ws.cell(row=row, column=3, value="")  # Filename
        ws.cell(row=row, column=4, value="")  # Size
        ws.cell(row=row, column=5, value=link_info["date"].strftime("%Y-%m-%d %H:%M:%S"))

    wb.save(filepath)
    wb.close()
    logger.info(f"Saved {len(links)} links to {filepath}")


async def extract_from_channel(
    api_id: int,
    api_hash: str,
    channel: str,
    limit: int = 100,
    output_file: str = "links.xlsx",
):
    """Connect to Telegram and extract TeraBox links from a channel/group."""

    channel = normalize_channel(channel)
    existing_links = load_existing_links(output_file)

    print(f"\n{'='*60}")
    print(f"  Telegram TeraBox Link Extractor")
    print(f"  Channel: {channel}")
    print(f"  Scanning: last {limit} messages")
    print(f"  Output:  {output_file}")
    if existing_links:
        print(f"  Existing: {len(existing_links)} links (will skip duplicates)")
    print(f"{'='*60}\n")

    async with TelegramClient(SESSION_FILE, api_id, api_hash) as client:
        # This will prompt for phone + OTP on first run
        if not await client.is_user_authorized():
            logger.info("First-time login — you'll be prompted for phone number and OTP code.")

        await client.start()
        logger.info("Connected to Telegram.")

        # Resolve the channel/group
        try:
            entity = await client.get_entity(channel)
            title = getattr(entity, "title", channel)
            logger.info(f"Found: {title}")
        except Exception as e:
            logger.error(f"Could not find channel '{channel}': {e}")
            logger.info("Make sure you are a member of the channel/group.")
            return []

        # Scan messages
        new_links = []
        scanned = 0
        duplicates = 0

        async for message in client.iter_messages(entity, limit=limit):
            scanned += 1
            urls = extract_urls_from_message(message)

            for url in urls:
                if url in existing_links:
                    duplicates += 1
                    continue
                existing_links.add(url)
                new_links.append({
                    "url": url,
                    "date": message.date.replace(tzinfo=None) if message.date else datetime.now(),
                    "message_id": message.id,
                })

            if scanned % 50 == 0:
                logger.info(f"  Scanned {scanned} messages, found {len(new_links)} new links...")

    logger.info(f"Scan complete: {scanned} messages scanned")
    logger.info(f"  New links:    {len(new_links)}")
    logger.info(f"  Duplicates:   {duplicates}")

    if new_links:
        save_to_excel(new_links, output_file, append=os.path.exists(output_file))
        print(f"\n  {len(new_links)} new TeraBox links saved to {output_file}")
    else:
        print(f"\n  No new TeraBox links found in the last {limit} messages.")

    return new_links


def trigger_batch_download(input_file: str, output_dir: str):
    """Trigger batch_downloader.py as a subprocess."""
    import subprocess

    script = os.path.join(PROJECT_DIR, "batch_downloader.py")
    python = sys.executable

    print(f"\n{'='*60}")
    print(f"  Triggering Batch Downloader...")
    print(f"{'='*60}\n")

    result = subprocess.run(
        [python, script, "--input", input_file, "--output-dir", output_dir],
        cwd=PROJECT_DIR,
    )
    return result.returncode


def main():
    parser = argparse.ArgumentParser(
        description="Extract TeraBox links from a Telegram channel/group and download them."
    )
    parser.add_argument(
        "--channel", "-c",
        required=True,
        help="Telegram channel/group: @username, t.me/channel, or invite link",
    )
    parser.add_argument(
        "--limit", "-l",
        type=int,
        default=100,
        help="Number of recent messages to scan (default: 100)",
    )
    parser.add_argument(
        "--output", "-o",
        default="links.xlsx",
        help="Output Excel file path (default: links.xlsx)",
    )
    parser.add_argument(
        "--download-dir", "-d",
        default=os.path.join(PROJECT_DIR, "downloads"),
        help="Download directory (default: ./downloads/)",
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Only extract links to Excel, don't trigger downloads",
    )

    args = parser.parse_args()

    # Step 1: Get Telegram credentials
    api_id, api_hash = get_telegram_credentials()

    # Step 2: Extract links from Telegram
    new_links = asyncio.run(
        extract_from_channel(
            api_id=api_id,
            api_hash=api_hash,
            channel=args.channel,
            limit=args.limit,
            output_file=args.output,
        )
    )

    # Step 3: Trigger batch downloader (if links found and not --no-download)
    if new_links and not args.no_download:
        trigger_batch_download(args.output, args.download_dir)
    elif not new_links:
        print("\n  No new links to download.")
    else:
        print(f"\n  Links saved. Run manually: python batch_downloader.py --input {args.output}")


if __name__ == "__main__":
    main()
