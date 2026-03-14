"""
Batch TeraBox Downloader — reads links from Excel, downloads all media files,
and writes status (Success/Failed) back into the Excel file.

Supports multiprocessing: runs multiple downloads in parallel.

Excel format:
  Column A: TeraBox Link
  Column B: Status (Success / Failed: reason)
  Column C: Filename
  Column D: Size (MB)
  Column E: Date (download timestamp)

Usage:
  python batch_downloader.py                          # default: links.xlsx, 2 workers
  python batch_downloader.py --input mylinks.xlsx     # custom Excel file
  python batch_downloader.py --workers 3              # 3 parallel downloads
  python batch_downloader.py --input links.xlsx --output-dir D:\\Videos --workers 4
"""

import argparse
import os
import sys
import time
import logging
from datetime import datetime
from multiprocessing import Pool, cpu_count

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Import streamer_v2 functions ─────────────────────────────────────────────
# Add project root to path so we can import sibling modules
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import SUPPORTED_DOMAINS

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] [PID %(process)d] %(message)s",
)
logger = logging.getLogger(__name__)

# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADERS = ["Link", "Status", "Filename", "Size (MB)", "Date"]


def is_terabox_url(url: str) -> bool:
    """Check if a URL belongs to a known TeraBox domain."""
    if not url or not url.startswith("http"):
        return False
    try:
        from urllib.parse import urlparse
        host = urlparse(url.strip()).hostname
        return host in SUPPORTED_DOMAINS if host else False
    except Exception:
        return False


def create_sample_excel(filepath: str):
    """Create a sample Excel file with headers and example link."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TeraBox Links"

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 22

    # Example row
    ws.cell(row=2, column=1, value="https://terasharefile.com/s/EXAMPLE_LINK_HERE")

    wb.save(filepath)
    logger.info(f"Created sample Excel: {filepath}")
    logger.info("Add your TeraBox links in column A, then run again.")


def load_excel(filepath: str):
    """Load workbook and return (workbook, worksheet)."""
    wb = load_workbook(filepath)
    ws = wb.active

    # Ensure headers exist in row 1
    if ws.cell(row=1, column=1).value != "Link":
        # Insert headers if missing
        ws.insert_rows(1)
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    return wb, ws


def _download_one(args: tuple) -> dict:
    """
    Worker function for multiprocessing — downloads a single TeraBox link.
    Must be a top-level function (picklable).
    Returns a result dict: {row_idx, link, status, filename, size_mb, timestamp, error}
    """
    row_idx, link, output_dir, max_retries = args

    # Import inside worker so each process gets its own module instance
    from streamer_v2 import get_download_link, download_file

    result = {
        "row_idx": row_idx,
        "link": link,
        "status": "Failed",
        "filename": None,
        "size_mb": None,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "error": None,
    }

    try:
        logger.info(f"[Row {row_idx}] Getting download link: {link[:60]}...")
        api_result = get_download_link(link, max_retries=max_retries)

        if "download_url" not in api_result:
            error_msg = api_result.get("error", "No download URL returned")
            raise RuntimeError(error_msg)

        dl_url = api_result["download_url"]

        logger.info(f"[Row {row_idx}] Downloading file...")
        filepath = download_file(dl_url, output_dir=output_dir)
        filename = os.path.basename(filepath)
        actual_size = os.path.getsize(filepath) / (1024 * 1024)

        result["status"] = "Success"
        result["filename"] = filename
        result["size_mb"] = round(actual_size, 2)
        result["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logger.info(f"[Row {row_idx}] SUCCESS: {filename} ({actual_size:.1f} MB)")

    except Exception as e:
        error_msg = str(e)[:100]
        result["error"] = error_msg
        result["status"] = f"Failed: {error_msg}"
        result["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logger.error(f"[Row {row_idx}] FAILED: {error_msg}")

    return result


def process_downloads(input_file: str, output_dir: str, max_retries: int = 3, workers: int = 2):
    """Main batch download: read Excel, download links in parallel, write status."""

    if not os.path.exists(input_file):
        logger.warning(f"Excel file not found: {input_file}")
        create_sample_excel(input_file)
        return

    wb, ws = load_excel(input_file)
    total_rows = ws.max_row - 1  # exclude header
    if total_rows <= 0:
        logger.warning("No links found in Excel file.")
        wb.close()
        return

    # Ensure column widths are set
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 22

    # ── Collect tasks ────────────────────────────────────────────────────────
    tasks = []       # (row_idx, link, output_dir, max_retries)
    skipped = 0
    invalid = 0

    for row_idx in range(2, ws.max_row + 1):
        link = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=2).value

        # Skip empty rows
        if not link or not str(link).strip():
            continue

        link = str(link).strip()

        # Skip already successful downloads
        if status and str(status).strip().lower() == "success":
            logger.info(f"Skipping (already downloaded): {link[:60]}...")
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = SKIP_FILL
            skipped += 1
            continue

        # Validate URL
        if not is_terabox_url(link):
            logger.warning(f"Invalid URL: {link[:60]}...")
            ws.cell(row=row_idx, column=2).value = "Failed: Invalid TeraBox URL"
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = FAIL_FILL
            invalid += 1
            continue

        tasks.append((row_idx, link, output_dir, max_retries))

    # Save any skip/invalid changes before starting pool
    wb.save(input_file)

    if not tasks:
        logger.info("No new links to download.")
        wb.close()
        return

    # Clamp workers to number of tasks
    actual_workers = min(workers, len(tasks))

    print(f"\n{'='*60}")
    print(f"  Batch TeraBox Downloader (Multiprocessing)")
    print(f"  Input:   {input_file}")
    print(f"  Output:  {output_dir}")
    print(f"  Links:   {len(tasks)} to download, {skipped} skipped, {invalid} invalid")
    print(f"  Workers: {actual_workers}")
    print(f"{'='*60}\n")

    # ── Run downloads in parallel ────────────────────────────────────────────
    succeeded = 0
    failed = invalid  # count invalid URLs as failed

    with Pool(processes=actual_workers) as pool:
        results = pool.map(_download_one, tasks)

    # ── Write results back to Excel ──────────────────────────────────────────
    for res in results:
        row_idx = res["row_idx"]

        ws.cell(row=row_idx, column=2).value = res["status"]
        ws.cell(row=row_idx, column=3).value = res["filename"]
        ws.cell(row=row_idx, column=4).value = res["size_mb"]
        ws.cell(row=row_idx, column=5).value = res["timestamp"]

        if res["status"] == "Success":
            fill = SUCCESS_FILL
            succeeded += 1
        else:
            fill = FAIL_FILL
            failed += 1

        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = fill

    wb.save(input_file)
    wb.close()

    # ── Summary ──────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  BATCH DOWNLOAD COMPLETE")
    print(f"  Succeeded: {succeeded}")
    print(f"  Failed:    {failed}")
    print(f"  Skipped:   {skipped}")
    print(f"  Total:     {total_rows}")
    print(f"  Workers:   {actual_workers}")
    print(f"  Excel:     {os.path.abspath(input_file)}")
    print(f"{'='*60}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Batch download TeraBox files from an Excel spreadsheet."
    )
    parser.add_argument(
        "--input", "-i",
        default="links.xlsx",
        help="Path to the Excel file with TeraBox links (default: links.xlsx)",
    )
    parser.add_argument(
        "--output-dir", "-o",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads"),
        help="Directory to save downloaded files (default: ./downloads/)",
    )
    parser.add_argument(
        "--retries", "-r",
        type=int,
        default=3,
        help="Max retries per link (default: 3)",
    )
    parser.add_argument(
        "--workers", "-w",
        type=int,
        default=2,
        help="Number of parallel downloads (default: 2, max recommended: 4)",
    )

    args = parser.parse_args()
    process_downloads(args.input, args.output_dir, max_retries=args.retries, workers=args.workers)


if __name__ == "__main__":
    main()
